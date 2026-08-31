"""
services/reporting.py

General-purpose report export — Dataset, CSV/XLSX/Markdown rendering,
and the operator's required filename format. Deliberately NOT under
services/assistant/: the operator's Excel-export requirement applies
to every data-bearing page (Flight Log, Crew Data, Roster, ...), not
just the OCC assistant's seven report templates. Putting this here
means a page's own "export currently filtered data" button and the
assistant's report functions share one implementation instead of two
— the assistant's services/assistant/reports.py imports from here,
it doesn't own this.

Dataset/AirlineIdentity below are taken from the assistant bundle's
services/assistant/models.py (received 2026-08-01) — that file also
defines ToolResult/QueryRequest/AuditEvent/AssistantAnswer, which look
like they belong to a different, LLM-tool-calling architecture
(provider_mode, citations, tools_used) than the deterministic,
no-LLM approach already decided and built (services/assistant/
query_parser.py). Only Dataset and AirlineIdentity are used here;
the rest of that file is not brought in.
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO, StringIO
from typing import Any, Sequence


@dataclass(frozen=True)
class AirlineIdentity:
    code: str
    name: str
    timezone: str = "UTC"


@dataclass(frozen=True)
class Dataset:
    """Structured rows for Streamlit tables and safe exports."""

    name: str
    title: str
    headers: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    notes: tuple[str, ...] = ()

    @classmethod
    def build(
        cls,
        *,
        name: str,
        title: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        notes: Sequence[str] = (),
    ) -> "Dataset":
        width = len(headers)
        normalized_rows = tuple(tuple(row) for row in rows)
        if any(len(row) != width for row in normalized_rows):
            raise ValueError("Every dataset row must match the header width")
        return cls(
            name=name,
            title=title,
            headers=tuple(headers),
            rows=normalized_rows,
            notes=tuple(notes),
        )


# Single-tenant default for the current deployment. configs/airlines/AEAGLE/
# is the eventual home for this if/when a second airline is ever onboarded
# (HANDOVER.md notes it's still empty) — a plain constant is deliberately
# not over-built into a config-loading mechanism for one airline.
AIR_EAGLE = AirlineIdentity(code="AEAGLE", name="AirEagle", timezone="Asia/Karachi")


def _cell(value: object) -> str:
    return "" if value is None else str(value)


def dataset_to_markdown(dataset: Dataset) -> str:
    def escape(value: object) -> str:
        return _cell(value).replace("|", r"\|").replace("\n", " ")

    header = "| " + " | ".join(map(escape, dataset.headers)) + " |"
    separator = "| " + " | ".join("---" for _ in dataset.headers) + " |"
    rows = [
        "| " + " | ".join(map(escape, row)) + " |"
        for row in dataset.rows
    ]
    table = "\n".join([header, separator, *rows]) if rows else "_No matching records found._"
    if not dataset.notes:
        return table
    notes_block = "\n".join(f"- {note}" for note in dataset.notes)
    return f"{table}\n\n**Notes:**\n{notes_block}"


def dataset_to_csv(dataset: Dataset) -> bytes:
    """Timestamps go out in Python's default ISO-ish form, NOT the
    `25 Aug 2003z` the screen uses.

    THE DIVERGENCE IS DELIBERATE (operator decision, 2026-08-31) and is
    not an inconsistency to tidy up. A CSV is a machine-readable
    regulatory artefact that outlives the session it was exported from:
    `2026-08-25 20:03:35` sorts correctly as text, parses in every tool
    without a custom format, and carries the YEAR — which `2003z` does
    not. The screen format is optimised for a controller reading a
    column at a glance; this one is optimised for whatever opens the
    file in two years.

    Anyone tempted to make the export "match the screen" is proposing to
    drop the year from a compliance record. See
    display_labels.utc_stamp() for the other side of the same decision.
    """
    stream = StringIO(newline="")
    writer = csv.writer(stream)
    writer.writerow(dataset.headers)
    writer.writerows(dataset.rows)
    if dataset.notes:
        writer.writerow([])
        writer.writerow(["Notes:"])
        for note in dataset.notes:
            writer.writerow([note])
    return stream.getvalue().encode("utf-8-sig")


def dataset_to_xlsx(dataset: Dataset) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Report"
    sheet.append(list(dataset.headers))
    for row in dataset.rows:
        sheet.append([_cell(value) for value in row])
    fill = PatternFill("solid", fgColor="0B5FA5")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(max(len(_cell(cell.value)) for cell in column) + 2, 45)
        sheet.column_dimensions[column[0].column_letter].width = width

    # Notes get their own sheet, deliberately not mixed into the data
    # rows — a note like "FDP is per-duty, do not sum across rows"
    # (see services/assistant/reports.py's crew_duty_history) needs to
    # survive being seen, not get lost as an oddly-shaped extra row in
    # a sheet someone's about to pivot/filter/sum.
    if dataset.notes:
        notes_sheet = workbook.create_sheet("Notes")
        notes_sheet.append(["Note"])
        notes_sheet["A1"].font = Font(bold=True)
        for note in dataset.notes:
            notes_sheet.append([note])
        notes_sheet.column_dimensions["A"].width = 100

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def report_filename(
    dataset: Dataset,
    airline: AirlineIdentity,
    extension: str,
    now: datetime | None = None,
) -> str:
    """
    AirEagle_FlightLog_24-07-2026_1735UTC.xlsx — the operator's
    required format: {airline.name}_{dataset.name}_DD-MM-YYYY_HHMMUTC.

    dataset.name must already be the exact PascalCase segment
    ("FlightLog", not "Flight Log") — safe() below strips anything
    outside [A-Za-z0-9_-], which would turn a space into an
    underscore rather than removing it, so a human-readable label
    would NOT produce the required format on its own.

    airline.name (e.g. "AirEagle") is a display string, deliberately
    separate from airline.code (e.g. "AEAGLE", used in
    services/audit_service.py's audit trail) — the two serve
    different purposes and are not derived from each other.

    safe() is still applied to airline.name: a no-op for "AirEagle"
    (nothing in it needs stripping) but a real backstop if a future
    airline's display name ever contains something the filename
    format can't carry.
    """
    timestamp = now or datetime.now(timezone.utc)
    safe = lambda value: re.sub(r"[^A-Za-z0-9_-]+", "_", value).strip("_")
    return (
        f"{safe(airline.name)}_{safe(dataset.name)}_"
        f"{timestamp:%d-%m-%Y_%H%MUTC}.{extension.lstrip('.')}"
    )
