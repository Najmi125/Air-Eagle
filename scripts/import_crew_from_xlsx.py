"""
scripts/import_crew_from_xlsx.py

Imports crew records from a filled-in copy of the crew data
collection template into the crew table, via crew_service.add_crew()
— never raw SQL, same rule as everywhere else in this repo.

This is NOT a one-off script for the 2026-07-21 data drop — it's
meant to be reused for every future batch (corrected loadmasters,
additional pilots, etc.), so it does real validation rather than
blindly trusting the spreadsheet:

  - Flags any row where a text field (Role/DOB/Nationality/Base/
    Ph No/Email/License No) actually contains a datetime value
    instead of text. That's not a typo — it's a sign the columns
    are misaligned relative to the template, exactly what happened
    with the 2026-07-21 Loadmaster section (rows filled in against a
    different header row that got pasted into the data area). Rows
    like this are SKIPPED, not guessed at.

  - Flags any date field with an implausible year (before 2020 or
    after 2100) rather than importing it silently. This does NOT
    auto-correct — seeing "1930" instead of "2030" needs a human
    decision (see KNOWN_CORRECTIONS below for how that decision gets
    recorded, not guessed by the script).

Usage:
    python scripts/import_crew_from_xlsx.py path/to/file.xlsx [--dry-run]
"""
import sys
import argparse
import datetime as dt
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from services import crew_service

# Maps the operator's actual column headers (as of the 2026-07-21
# data drop, after migration 007 reconciled SIM/Route Check/IR) to
# crew_service field names. If a future batch uses different header
# text, update this map — don't guess at a fuzzy match.
HEADER_MAP = {
    "ID": "operator_staff_id",
    "Name": "name",
    "Role": "role",
    "DOB": "date_of_birth",
    "Nationality": "nationality",
    "Base": "base",
    "Ph No": "phone",
    "Email": "email",
    "License No": "license_no",
    "License Exp": "license_expiry",
    "Medical Exp": "medical_expiry",
    "IR": "ir_expiry",
    "Type Rating Exp": "type_rating_expiry",
    "SIM Exp": "sim_expiry",
    "Route Check Exp": "route_check_expiry",
    "SEP Exp": "sep_expiry",
    "CRM Exp": "crm_expiry",
    "DG Exp": "dg_expiry",
    "Contract Exp": "contract_expiry",
    "Remarks": "remarks",
}

TEXT_FIELDS = {"role", "nationality", "base", "phone", "email", "license_no"}
DATE_FIELDS = {
    "date_of_birth", "license_expiry", "medical_expiry", "ir_expiry",
    "type_rating_expiry", "sim_expiry", "route_check_expiry", "sep_expiry",
    "crm_expiry", "dg_expiry", "contract_expiry",
}
# Subset of DATE_FIELDS subject to the plausible-year check below.
# date_of_birth is deliberately excluded — a birth year like 1960 is
# completely normal and would wrongly trip a check tuned for expiry
# dates, which should always be recent/near-future.
EXPIRY_FIELDS = DATE_FIELDS - {"date_of_birth"}
PLAUSIBLE_YEAR_RANGE = (2020, 2100)

# Explicit, reviewed corrections for specific known bad values in a
# specific data drop — NOT a general auto-correction mechanism. Each
# entry here should trace back to an actual confirmed decision, not
# a guess. Keyed by (row_number_in_sheet, field_name), value is
# {"expected_name": ..., "value": ...}.
#
# The expected_name cross-check is the actual safeguard here, not
# just documentation: a correction only applies if the row's real
# Name matches who it was reviewed for. Row number alone was proven
# unsafe during development — synthetic test rows also start at row
# 3 and silently inherited WAQAR's real correction before this was
# fixed. Matching by name too means even if rows get reordered,
# inserted, or a different file reuses row 3 for someone else
# entirely, the correction won't silently misapply — it'll be
# reported as skipped instead (see the name-mismatch warning below).
#
# Still clear these entries once a data drop's import is confirmed
# correct, before the next batch — this reduces the blast radius of
# a collision, it doesn't make stale entries a non-issue.
KNOWN_CORRECTIONS = {
    # 2026-07-21 drop: three License Exp dates read 01-Jun/Feb-1930,
    # confirmed by the user to be 2030 (source system likely dropped
    # the century on export).
    (3, "license_expiry"): {"expected_name": "MUHAMMAD WAQAR", "value": dt.date(2030, 6, 1)},
    (6, "license_expiry"): {"expected_name": "TAHIR MAHMOOD RAJA", "value": dt.date(2030, 2, 1)},
    (11, "license_expiry"): {"expected_name": "MUHAMMAD SHAHBAZ", "value": dt.date(2030, 6, 1)},
}


def _clean_value(row_num, field, value, row_name=None):
    if value is None:
        return None
    if isinstance(value, str) and value.strip() in ("-", ""):
        return None

    if (row_num, field) in KNOWN_CORRECTIONS:
        correction = KNOWN_CORRECTIONS[(row_num, field)]
        expected_name = correction["expected_name"]
        actual_name = (row_name or "").strip().upper()
        if actual_name == expected_name.strip().upper():
            return correction["value"]
        # Row number matched but the name didn't — this is exactly
        # the collision the review flagged. Don't apply a correction
        # meant for someone else; fall through to normal handling
        # (which will likely flag this as a suspect date instead,
        # correctly surfacing it for review rather than silently
        # "fixing" it wrong).

    if field in EXPIRY_FIELDS and isinstance(value, (dt.date, dt.datetime)):
        year = value.year
        if not (PLAUSIBLE_YEAR_RANGE[0] <= year <= PLAUSIBLE_YEAR_RANGE[1]):
            return "SUSPECT_DATE"  # sentinel, caller decides what to do

    if field in DATE_FIELDS and isinstance(value, (dt.date, dt.datetime)):
        return value.date() if isinstance(value, dt.datetime) else value

    return value


def read_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Crew Data"]
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]

    results = {"imported": [], "skipped_misaligned": [], "skipped_suspect_date": [], "skipped_empty": []}

    for row_num in range(3, ws.max_row + 1):
        raw = {headers[c - 1]: ws.cell(row=row_num, column=c).value for c in range(1, len(headers) + 1)}
        if all(v is None for v in raw.values()):
            continue
        if "(EXAMPLE)" in str(raw.get("Name", "")):
            continue

        row_name = raw.get("Name")
        record = {}
        misaligned = False
        suspect = False

        for header, value in raw.items():
            field = HEADER_MAP.get(header)
            if field is None:
                continue
            cleaned = _clean_value(row_num, field, value, row_name=row_name)

            # Symmetric check: a text field should never hold a date
            # (e.g. Base='2027-04-01'), and a date field should never
            # hold arbitrary text once "-"/empty are already handled
            # (e.g. DOB='SEP'). Either direction is the same signal —
            # this row was filled in against different columns than
            # this template actually has.
            if field in TEXT_FIELDS and isinstance(value, (dt.date, dt.datetime)):
                misaligned = True
                continue
            if field in DATE_FIELDS and cleaned is not None and cleaned != "SUSPECT_DATE" \
                    and not isinstance(cleaned, (dt.date, dt.datetime)):
                misaligned = True
                continue

            if cleaned == "SUSPECT_DATE":
                suspect = True
                continue

            record[field] = cleaned

        if misaligned:
            results["skipped_misaligned"].append((row_num, raw.get("Name")))
            continue
        if suspect:
            results["skipped_suspect_date"].append((row_num, raw.get("Name")))
            continue
        if not record.get("name") or not record.get("role"):
            results["skipped_empty"].append((row_num, raw.get("Name")))
            continue

        results["imported"].append((row_num, record))

    return results


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    results = read_rows(args.path)

    print(f"Rows ready to import: {len(results['imported'])}")
    print(f"Rows skipped (misaligned columns): {len(results['skipped_misaligned'])}")
    for row_num, name in results["skipped_misaligned"]:
        print(f"  row {row_num}: {name}")
    print(f"Rows skipped (suspect date, needs a KNOWN_CORRECTIONS entry): {len(results['skipped_suspect_date'])}")
    for row_num, name in results["skipped_suspect_date"]:
        print(f"  row {row_num}: {name}")
    print(f"Rows skipped (empty/missing name or role): {len(results['skipped_empty'])}")

    if args.dry_run:
        print("\n--dry-run: nothing written.")
        return

    print()
    for row_num, record in results["imported"]:
        crew_id = crew_service.add_crew(record, app_user="import_script")
        print(f"  row {row_num}: {record['name']} -> {crew_id}")


if __name__ == "__main__":
    main()
