"""
tests/test_import_crew_script.py

Builds small synthetic workbooks matching the template structure to
test scripts/import_crew_from_xlsx.py's validation logic in
isolation, without depending on the real operator file (which won't
always be available, and shouldn't be committed to the repo anyway).
"""
import sys
import datetime as dt
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
import pytest

from scripts.import_crew_from_xlsx import read_rows, HEADER_MAP
import scripts.import_crew_from_xlsx as import_mod

HEADERS = list(HEADER_MAP.keys())


@pytest.fixture(autouse=True)
def _isolate_known_corrections(monkeypatch):
    """
    Every test workbook here starts its data at row 3 (openpyxl's
    natural row numbering for a single-row list) — which collides
    with the real, hardcoded KNOWN_CORRECTIONS entries for the actual
    2026-07-21 operator data drop (WAQAR's real row 3). Without this,
    every test using the default license_expiry would silently get
    overwritten to WAQAR's correction value regardless of what the
    test actually set, since the correction lookup matches by row
    number alone, not by value.

    Isolating this dict to empty by default means each test is
    testing the general mechanism, not accidentally interacting with
    today's specific reviewed corrections. The one test that verifies
    the correction mechanism itself adds its own entry locally.
    """
    monkeypatch.setattr(import_mod, "KNOWN_CORRECTIONS", {})


def _build_workbook(rows, tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crew Data"
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    for r, row_data in enumerate(rows, start=3):
        for c, h in enumerate(HEADERS, start=1):
            ws.cell(row=r, column=c, value=row_data.get(h))
    path = tmp_path / "test_crew.xlsx"
    wb.save(path)
    return path


def _clean_row(**overrides):
    base = {
        "ID": "AE-01", "Name": "Test Person", "Role": "CPT",
        "DOB": dt.datetime(1980, 1, 1), "Nationality": "Pakistani", "Base": "KHI",
        "Ph No": "123", "Email": "test@example.com", "License No": "ATPL-1",
        "License Exp": dt.datetime(2029, 1, 1), "Medical Exp": dt.datetime(2026, 12, 1),
    }
    base.update(overrides)
    return base


def test_clean_row_is_imported(tmp_path):
    path = _build_workbook([_clean_row()], tmp_path)
    results = read_rows(path)
    assert len(results["imported"]) == 1
    assert results["skipped_misaligned"] == []


def test_dob_as_real_date_does_not_trigger_false_misalignment(tmp_path):
    """Regression test for the actual bug found during the 2026-07-21
    import: date_of_birth was originally miscategorized as a text
    field, causing every clean row's real DOB to be falsely flagged
    as misaligned."""
    path = _build_workbook([_clean_row(DOB=dt.datetime(1960, 6, 16))], tmp_path)
    results = read_rows(path)
    assert len(results["imported"]) == 1
    row_num, record = results["imported"][0]
    assert record["date_of_birth"] == dt.date(1960, 6, 16)


def test_date_in_text_field_is_caught_as_misaligned(tmp_path):
    """A date sitting in a field that should be text (e.g. Base) is
    a sign of column misalignment, not real data."""
    path = _build_workbook([_clean_row(Base=dt.datetime(2027, 4, 1))], tmp_path)
    results = read_rows(path)
    assert results["imported"] == []
    assert len(results["skipped_misaligned"]) == 1


def test_text_in_date_field_is_caught_as_misaligned(tmp_path):
    """The mirror case — the actual bug in the real 2026-07-21 data:
    a stray header label ('SEP') sitting in DOB, a date field."""
    path = _build_workbook([_clean_row(DOB="SEP")], tmp_path)
    results = read_rows(path)
    assert results["imported"] == []
    assert len(results["skipped_misaligned"]) == 1


def test_dash_and_empty_string_normalize_to_none(tmp_path):
    path = _build_workbook([_clean_row(ID="-", Nationality="")], tmp_path)
    results = read_rows(path)
    assert len(results["imported"]) == 1
    row_num, record = results["imported"][0]
    assert record["operator_staff_id"] is None
    assert record.get("nationality") is None


def test_implausible_expiry_year_is_flagged_not_imported(tmp_path):
    """An uncorrected implausible date (e.g. year 1930) must be
    skipped and reported, not silently imported as-is or guessed at."""
    path = _build_workbook([_clean_row(**{"License Exp": dt.datetime(1930, 6, 1)})], tmp_path)
    results = read_rows(path)
    assert results["imported"] == []
    assert len(results["skipped_suspect_date"]) == 1


def test_implausible_dob_year_is_not_flagged(tmp_path):
    """date_of_birth is deliberately excluded from the plausible-year
    check — 1960 is a completely normal birth year and must not be
    treated as suspect the way an expiry date would be."""
    path = _build_workbook([_clean_row(DOB=dt.datetime(1955, 3, 1))], tmp_path)
    results = read_rows(path)
    assert len(results["imported"]) == 1
    assert results["skipped_suspect_date"] == []


def test_known_correction_applies_when_name_matches(tmp_path):
    """A row with a KNOWN_CORRECTIONS entry should import cleanly
    with the corrected value when the row's actual name matches who
    the correction was reviewed for."""
    path = _build_workbook([_clean_row(**{"License Exp": dt.datetime(1930, 6, 1)})], tmp_path)

    # The autouse fixture already isolated KNOWN_CORRECTIONS to an
    # empty dict for this test — add the one entry this test needs.
    import_mod.KNOWN_CORRECTIONS[(3, "license_expiry")] = {
        "expected_name": "Test Person", "value": dt.date(2030, 6, 1),
    }

    results = read_rows(path)
    assert len(results["imported"]) == 1
    row_num, record = results["imported"][0]
    assert record["license_expiry"] == dt.date(2030, 6, 1)


def test_known_correction_does_not_apply_when_name_mismatches(tmp_path):
    """The actual safeguard the review asked for: a correction keyed
    to row 3 must NOT silently apply to a different person who
    happens to also be at row 3 in some other file. Row-number match
    alone is not enough — falls through to the normal suspect-date
    handling instead of misapplying someone else's correction."""
    path = _build_workbook([_clean_row(
        Name="A Completely Different Person",
        **{"License Exp": dt.datetime(1930, 6, 1)},
    )], tmp_path)

    import_mod.KNOWN_CORRECTIONS[(3, "license_expiry")] = {
        "expected_name": "Test Person", "value": dt.date(2030, 6, 1),
    }

    results = read_rows(path)
    assert results["imported"] == []
    assert len(results["skipped_suspect_date"]) == 1


def test_example_row_is_skipped(tmp_path):
    path = _build_workbook([_clean_row(Name="Someone (EXAMPLE)")], tmp_path)
    results = read_rows(path)
    assert results["imported"] == []


def test_missing_name_or_role_is_skipped(tmp_path):
    path = _build_workbook([_clean_row(Name=None)], tmp_path)
    results = read_rows(path)
    assert results["imported"] == []
    assert len(results["skipped_empty"]) == 1


def test_multiple_rows_mixed_outcomes(tmp_path):
    """Sanity check that one bad row doesn't affect processing of the
    others in the same file."""
    rows = [
        _clean_row(Name="Good One"),
        _clean_row(Name="Bad Alignment", Base=dt.datetime(2027, 1, 1)),
        _clean_row(Name="Good Two"),
    ]
    path = _build_workbook(rows, tmp_path)
    results = read_rows(path)
    assert len(results["imported"]) == 2
    assert len(results["skipped_misaligned"]) == 1
