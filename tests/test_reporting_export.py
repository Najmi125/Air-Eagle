"""
tests/test_reporting_export.py

Pure logic — no database needed. Covers services/reporting.py:
Dataset construction/validation, CSV/XLSX/Markdown rendering, and the
operator's required export filename format
(AirEagle_[PageName]_DD-MM-YYYY_HHMMUTC.xlsx).
"""
import csv
import datetime as dt
import sys
from io import BytesIO, StringIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pytest
from openpyxl import load_workbook

from services.reporting import (
    AIR_EAGLE,
    AirlineIdentity,
    Dataset,
    dataset_to_csv,
    dataset_to_markdown,
    dataset_to_xlsx,
    report_filename,
)


def _sample_dataset(notes=()):
    return Dataset.build(
        name="FlightLog",
        title="Flight Log",
        headers=["flight_id", "flight_no", "status"],
        rows=[[1, "EPE 786", "OPERATED"], [2, None, "CANCELLED"]],
        notes=notes,
    )


def test_dataset_build_normalizes_rows_to_tuples():
    ds = _sample_dataset()
    assert ds.headers == ("flight_id", "flight_no", "status")
    assert ds.rows == ((1, "EPE 786", "OPERATED"), (2, None, "CANCELLED"))


def test_dataset_build_rejects_mismatched_row_width():
    with pytest.raises(ValueError):
        Dataset.build(
            name="Bad", title="Bad",
            headers=["a", "b"],
            rows=[[1, 2], [1, 2, 3]],
        )


def test_dataset_build_accepts_empty_rows():
    ds = Dataset.build(name="Empty", title="Empty", headers=["a", "b"], rows=[])
    assert ds.rows == ()


def test_markdown_renders_header_and_rows():
    ds = _sample_dataset()
    md = dataset_to_markdown(ds)
    assert "| flight_id | flight_no | status |" in md
    assert "| 1 | EPE 786 | OPERATED |" in md
    assert "| 2 |  | CANCELLED |" in md  # None -> empty cell, not "None"


def test_markdown_empty_dataset_says_no_records():
    ds = Dataset.build(name="Empty", title="Empty", headers=["a"], rows=[])
    md = dataset_to_markdown(ds)
    assert md == "_No matching records found._"


def test_markdown_appends_notes_section_only_when_present():
    without_notes = dataset_to_markdown(_sample_dataset())
    assert "**Notes:**" not in without_notes

    with_notes = dataset_to_markdown(_sample_dataset(notes=["Do not sum fdp_hours across rows."]))
    assert "**Notes:**" in with_notes
    assert "- Do not sum fdp_hours across rows." in with_notes


def test_markdown_escapes_pipes_and_newlines_in_cells():
    ds = Dataset.build(
        name="Weird", title="Weird", headers=["note"],
        rows=[["line one | with pipe\nline two"]],
    )
    md = dataset_to_markdown(ds)
    assert r"line one \| with pipe line two" in md


def test_csv_round_trips_headers_and_rows():
    ds = _sample_dataset()
    raw = dataset_to_csv(ds)
    text = raw.decode("utf-8-sig")
    reader = csv.reader(StringIO(text))
    rows = list(reader)
    assert rows[0] == ["flight_id", "flight_no", "status"]
    assert rows[1] == ["1", "EPE 786", "OPERATED"]
    assert rows[2] == ["2", "", "CANCELLED"]  # None -> empty string, csv.writer default


def test_csv_appends_notes_after_blank_row():
    ds = _sample_dataset(notes=["Note one.", "Note two."])
    raw = dataset_to_csv(ds)
    rows = list(csv.reader(StringIO(raw.decode("utf-8-sig"))))
    assert rows[3] == []
    assert rows[4] == ["Notes:"]
    assert rows[5] == ["Note one."]
    assert rows[6] == ["Note two."]


def test_csv_no_notes_section_when_no_notes():
    raw = dataset_to_csv(_sample_dataset())
    rows = list(csv.reader(StringIO(raw.decode("utf-8-sig"))))
    assert len(rows) == 3  # header + 2 data rows, nothing else


def test_xlsx_data_sheet_matches_dataset():
    ds = _sample_dataset()
    raw = dataset_to_xlsx(ds)
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames[0] == "Report"
    sheet = wb["Report"]
    assert [c.value for c in sheet[1]] == ["flight_id", "flight_no", "status"]
    assert [c.value for c in sheet[2]] == ["1", "EPE 786", "OPERATED"]
    # openpyxl reads a written empty string back as None, not "" —
    # Excel itself doesn't distinguish "empty cell" from "empty string
    # cell" on round-trip.
    assert [c.value for c in sheet[3]] == ["2", None, "CANCELLED"]
    assert sheet.freeze_panes == "A2"


def test_xlsx_header_row_is_styled():
    raw = dataset_to_xlsx(_sample_dataset())
    wb = load_workbook(BytesIO(raw))
    header_cell = wb["Report"][1][0]
    assert header_cell.font.bold is True


def test_xlsx_notes_get_a_separate_sheet_only_when_present():
    without_notes = load_workbook(BytesIO(dataset_to_xlsx(_sample_dataset())))
    assert "Notes" not in without_notes.sheetnames

    with_notes = load_workbook(BytesIO(dataset_to_xlsx(
        _sample_dataset(notes=["FDP is per-duty, do not sum across rows."])
    )))
    assert "Notes" in with_notes.sheetnames
    notes_sheet = with_notes["Notes"]
    assert notes_sheet["A1"].value == "Note"
    assert notes_sheet["A2"].value == "FDP is per-duty, do not sum across rows."
    # data sheet itself stays clean — no extra rows appended for notes
    assert with_notes["Report"].max_row == 3


def test_report_filename_matches_operator_required_format():
    ds = Dataset.build(name="FlightLog", title="Flight Log", headers=["a"], rows=[])
    now = dt.datetime(2026, 7, 24, 17, 35, tzinfo=dt.timezone.utc)
    assert report_filename(ds, AIR_EAGLE, "xlsx", now=now) == "AirEagle_FlightLog_24-07-2026_1735UTC.xlsx"


def test_report_filename_strips_leading_dot_from_extension():
    ds = Dataset.build(name="FlightLog", title="Flight Log", headers=["a"], rows=[])
    now = dt.datetime(2026, 7, 24, 17, 35, tzinfo=dt.timezone.utc)
    assert report_filename(ds, AIR_EAGLE, ".csv", now=now).endswith(".csv")
    assert ".." not in report_filename(ds, AIR_EAGLE, ".csv", now=now)


def test_report_filename_sanitizes_unsafe_dataset_name():
    """dataset.name should already be exact PascalCase, but safe() is
    a backstop: a space collapses to underscore rather than vanishing
    (which would silently produce a filename that doesn't match the
    required format at all)."""
    ds = Dataset.build(name="Flight Log!", title="Flight Log", headers=["a"], rows=[])
    now = dt.datetime(2026, 7, 24, 17, 35, tzinfo=dt.timezone.utc)
    name = report_filename(ds, AIR_EAGLE, "xlsx", now=now)
    assert name == "AirEagle_Flight_Log_24-07-2026_1735UTC.xlsx"


def test_report_filename_sanitizes_unsafe_airline_name():
    ds = Dataset.build(name="FlightLog", title="Flight Log", headers=["a"], rows=[])
    weird_airline = AirlineIdentity(code="XX", name="Weird/Air Co.")
    now = dt.datetime(2026, 7, 24, 17, 35, tzinfo=dt.timezone.utc)
    name = report_filename(ds, weird_airline, "xlsx", now=now)
    assert name.startswith("Weird_Air_Co")


def test_report_filename_defaults_to_now_when_not_given():
    ds = Dataset.build(name="FlightLog", title="Flight Log", headers=["a"], rows=[])
    before = dt.datetime.now(dt.timezone.utc)
    name = report_filename(ds, AIR_EAGLE, "xlsx")
    after = dt.datetime.now(dt.timezone.utc)
    stamp = name.split("_")[-1]  # e.g. "24-07-2026" + ... actually last segment before ext
    # Just confirm it parses as today's date somewhere in the filename,
    # rather than asserting an exact string (a real "now" call is
    # inherently non-deterministic to the second).
    assert before.strftime("%d-%m-%Y") in name or after.strftime("%d-%m-%Y") in name
